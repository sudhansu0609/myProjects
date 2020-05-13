from openpyxl import Workbook, load_workbook, worksheet, workbook
from contextlib import closing
from fileinput import filename
import os.path
from openpyxl.xml.constants import MAX_ROW
import xlrd
from builtins import int

def find_Row_Values(loc,sheetName,rowNum):
    rowNum=int(rowNum)  
    wb = xlrd.open_workbook(loc) 
    sheet = wb.sheet_by_name(sheetName)
    myHeadingKeyList=[]
    myHeadingValueList=[]
    try:
        for i in range(sheet.ncols):
            curentKey=sheet.cell_value(0, i)
            myHeadingKeyList.append(curentKey)
        for i in range(sheet.ncols):
            curentKey=sheet.cell_value(rowNum,i)
            myHeadingValueList.append(curentKey)
            print(str(myHeadingKeyList[i])+'='+str(myHeadingValueList[i]))
    except IndexError as IndException:
        print("Input Index not found: "+str(IndException))
    return  dict(zip(myHeadingKeyList,myHeadingValueList))

def findColName(loc,rowNum):
    rowNum=int(rowNum)  
    wb = xlrd.open_workbook(loc) 
    sheet = wb.sheet_by_index(0)
    myHeadingKeyList=[]
    myHeadingValueList=[]
    try:
        for i in range(sheet.ncols):
            curentKey=sheet.cell_value(0, i)
            myHeadingKeyList.append(curentKey)
        for i in range(sheet.ncols):
            curentKey=sheet.cell_value(rowNum,i)
            myHeadingValueList.append(curentKey)
            print(str(myHeadingKeyList[i])+'='+str(myHeadingValueList[i]))
    except IndexError as IndException:
        print("Input Index not found: "+str(IndException))
    return  dict(zip(myHeadingKeyList,myHeadingValueList))
	
def create_excel_file(sheet_name, file_name):
    with closing(Workbook()) as wb:
        ws = wb.active
        ws.title = sheet_name
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        wb.save(file_name)
        
def add_new_sheet(sheet_name, file_name):
    with closing(load_workbook(filename=file_name)) as wb:
        wb.create_sheet(sheet_name)
        ws = wb[sheet_name]
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        wb.save(file_name)

def add_value(sheet_name, file_name, cell_cords, value):
    with closing(load_workbook(filename=file_name)) as wb:
        ws = wb[sheet_name]
        ws[cell_cords] = value
        wb.save(file_name)
        
def get_row_count(sheet_name, file_name):
    with closing(load_workbook(filename=file_name)) as wb:
        ws = wb[sheet_name]
        return ws.max_row
        
def append_value(sheet_name, file_name, c1_value, c2_value):
    with closing(load_workbook(filename=file_name)) as wb:
        ws = wb[sheet_name]
        ws.append([c1_value, c2_value])
        wb.save(file_name)
#name changed from append_value to append_value1 to remove duplicacy.change it back if required  		
def append_values(sheet_name, file_name, c1_value, c2_value, c3_value):
    with closing(load_workbook(filename=file_name)) as wb:
        ws = wb[sheet_name]
        ws.append([c1_value, c2_value, c3_value])
        wb.save(file_name)
		
def check_file(path):
    if os.path.isfile(path):
        return True
    else:
        return False
        
def find_model(sheet_name, file_name, value):
    with closing(load_workbook(filename=file_name)) as wb:
        ws = wb[sheet_name]
        for row in ws.iter_rows(max_row=ws.max_row, min_col=2):
            for cell in row:
                if (value == cell.value):
                    cr = 'A'+str(cell.row)
                    return ws[cr].value

        else:
            return 0
                
                
def find_value(sheet_name, file_name, value):
    with closing(load_workbook(filename=file_name)) as wb:
        ws = wb[sheet_name]
        for row in ws.iter_rows(max_row=ws.max_row, min_col=2):
            for cell in row:
                if (value == cell.value):

                    
                    return cell.row
                
        else:
            return 0            

            
            
def check_sheet(sheet_name, file_name):
    with closing(load_workbook(filename=file_name, read_only=True)) as wb:
        if sheet_name in wb.sheetnames:
            return True
        else:
            return False
        
def find_uuid(sheet_name, file_name):
    with closing(load_workbook(filename=file_name)) as wb:
        ws = wb[sheet_name]
        for row in ws.iter_rows(max_row=ws.max_row, min_col=2):
            for cell in row:
                if ('Distributed' == cell.value):
                    cr = 'C'+str(cell.row)
                    return ws[cr].value

                else:
                    return 0
			
def find_uuid_AAI(sheet_name, file_name):
    with closing(load_workbook(filename=file_name)) as wb:
        ws = wb[sheet_name]
        for row in ws.iter_rows(max_row=ws.max_row, min_col=2):
            for cell in row:
                if ('Distributed(MSO Verified)' == cell.value):
                    cr = 'C'+str(cell.row)

                    return ws[cr].value        
        else:
            return 0
        
def excel_rowColumnLookup(sheet_name, file_name, value):
    with closing(load_workbook(filename=file_name)) as wb:
        ws = wb[sheet_name]
        for row in ws.iter_rows(max_row=ws.max_row, min_col=2):
            for cell in row:
                if (value == cell.value):
                    cellAddress=dict()
                    cellAddress['row']=cell.row
                    cellAddress['column']=cell.column                    
                    return cellAddress
        else:
            return False

