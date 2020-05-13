# class pandasExcelManipulation(object):
# 
import os
import zipfile as zf
import openpyxl as op
from time import sleep
from openpyxl.worksheet.header_footer import _split_string
from os import listdir
from os.path import isfile ,join
import re
#import  Calculate_And_Validate_Ip as cIp


import shutil

class PreloadExcelManipulation(object):
    def __init__(self):
       pass

    def unzip_heatFile(self,heatfile):
        with zf.ZipFile(heatfile,"r") as zip_ref:
           # print(heatfile)
            fileName=heatfile.split("/")[-1].split(".")[0]
            #Extension=heatfile.split("/")[-1].split(".")[1]
            #print(fileName)
            
            extractedFolder="C://robot_framework//robotsolution//robot//assets//SDC//heatfiles//"+ fileName
            zip_ref.extractall(extractedFolder)
        return  extractedFolder
    
    
    def loop_over_heatFiles(self,heatfile,rolefile):
        path=self.unzip_heatFile(heatfile) 
        rolefile= join(rolefile,"NetworkRoles")
        unzipfiles=[f for f in listdir(path) if isfile(join(path, f)) and (not(join(path, f).endswith('.txt')))] 
        print(unzipfiles)
        #unzipfiles=
        fileList=[]
        for file in unzipfiles:
            fileList.append(join(path,file))
       # print(fileList)    
        uniqueRoles=[]
        for file in fileList:
           #  with open(file,"r") as uf,open(rolefile,"r") as rf:
             lines=[]
             roles=[]
             with open(file,"r") as uf, open(rolefile,'r') as rf:
                 for line in uf:
                     lines.append(line.strip("\n"))        
                 for role in rf:
                     roles.append(role.strip("\n")) 
                 for line1 in lines:
                     for role1 in roles:
                         if line1.find(role1+"_")!=-1 :
                             print(line1+"||                  "+role1 +"              ||||true") 
                             if role1 not in uniqueRoles:
                                uniqueRoles.append(role1) 
                           # print(role1)
                                #print(line1+"  ||||true") 
                     #else:
                          #print("false")
        print(uniqueRoles) 
        return uniqueRoles
#heatfile is gull path of heatfile
#role file is full path of role file
#vnfName is name of vnf whose heatfile is given 
#path is path till preload excels inwhich folder of vng name is to be created
#C://robot_framework//robotsolution//robot//assets//sdnc-portal//ManualPreloadFiles//Preload_Excels     
    def Update_Preload_Excel(self, excelPath, fileName, Sheetname, searchValue):
         excelPathFull=os.path.join(excelPath,fileName,searchValue)
         networkValues=searchValue.split("=")
         networkParameter=networkValues[0]
         networkParameterValue=networkValues[1]
         wb=op.load_workbook(excelPathFull,keep_vba='KEEP_VBA')
         ws=wb[Sheetname]
         maxRow=ws.max_row
        # maxCol=ws.max_col
         for row in ws.iter_rows():
             for cell in row:
                 if str(cell.value).strip()==networkParameter:
                    celRowNo=cell.row
                    celColNo=cell.column
                    print(ws.cell(celRowNo,celColNo+1).value)
                    writtenCell=ws.cell(celRowNo,celColNo+1)
                    ws.cell(celRowNo,celColNo+1).value=networkParameterValue
                   # writtenCell="sudhansu"
                    print(str(cell.value).strip())
                    print("value found")
         wb.save(excelPathFull)
         print(maxRow)  
    def Update_GeneralSheet(self,filePath,networkParameterValue,labname):
             excelPathFull=filePath                                     #os.path.join(excelPath,fileName,searchValue)
             Sheetname='General'
             wb=op.load_workbook(excelPathFull,read_only=False,keep_vba='KEEP_VBA')
             ws=wb[Sheetname]
             maxRow=ws.max_row
            # maxCol=ws.max_col
             #for networkParameter in 'network-name','network-type':
             for row in ws.iter_rows():
                 for cell in row:
                     if str(cell.value).strip()=="network-role":
                        celRowNo=cell.row
                        celColNo=cell.column
                        print(ws.cell(celRowNo,celColNo+1).value)
                        writtenCell=ws.cell(celRowNo,celColNo+1)
                        ws.cell(celRowNo,celColNo+1).value=networkParameterValue
                       # writtenCell="sudhansu"
                        print(str(cell.value).strip())
                        print("value found")
             for row in ws.iter_rows():
                 for cell in row:
                     if str(cell.value).strip()=="network-name":
                        celRowNo=cell.row
                        celColNo=cell.column
                        print(ws.cell(celRowNo,celColNo+1).value)
                        writtenCell=ws.cell(celRowNo,celColNo+1)
                        ws.cell(celRowNo,celColNo+1).value=labname+"_"+networkParameterValue
                       # writtenCell="sudhansu"
                        print(str(cell.value).strip())
                        print("value found")          
             wb.save(excelPathFull)   
            
             print(maxRow) 
    def Update_SubnetsSheet(self,filePath,networkParameterValue,labname,ipv6Address,ipV6Cidr,ipv4Address,ipV4Cidr):
             excelPathFull=filePath                                     #os.path.join(excelPath,fileName,searchValue)
             Sheetname='General'
             wb=op.load_workbook(excelPathFull,keep_vba='KEEP_VBA')
             ws=wb[Sheetname]
             maxRow=ws.max_row
            # maxCol=ws.max_col
             subnetNames=['start-address','dhcp-start','dhcp-end','gateway-address']
             for networkParameter in subnetNames:
                 for row in ws.iter_rows():
                     for cell in row:
                         if str(cell.value).strip()=="start-address":
                            celRowNo=cell.row
                            celColNo=cell.column
                            print(ws.cell(celRowNo,celColNo+1).value)
                            Ipv4_writtenCell=ws.cell(celRowNo+1,celColNo)
                            ws.cell(celRowNo+1,celColNo).value=cIp.get_ip4_Start_Address(ipv4Address,ipV4Cidr) #in networkParameterValue give start address from ip library
                            Ipv6_writtenCell=ws.cell(celRowNo+1,celColNo)
                            ws.cell(celRowNo+2,celColNo).value=cIp.calculate_ipv6_address_Start(ipv6Address,ipV6Cidr)   #in networkParameterValue give start address from ip library
                         elif str(cell.value).strip()=="dhcp-start":  
                                celRowNo=cell.row
                                celColNo=cell.column
                                print(ws.cell(celRowNo,celColNo+1).value)
                                Ipv4_writtenCell=ws.cell(celRowNo+1,celColNo)
                                ws.cell(celRowNo+1,celColNo).value=cIp.get_ip4_dhcp_start(ipv4Address,ipV4Cidr)   #in networkParameterValue give start address from ip library
                                Ipv6_writtenCell=ws.cell(celRowNo+1,celColNo)
                                ws.cell(celRowNo+2,celColNo).value=cIp.Calculate_IpV6_dhcp_Start(ipv6Address,ipV6Cidr)   #in networkParameterValue give start address from ip library
                         elif str(cell.value).strip()=="dhcp-end":  
                                celRowNo=cell.row
                                celColNo=cell.column
                                print(ws.cell(celRowNo,celColNo+1).value)
                                Ipv4_writtenCell=ws.cell(celRowNo+1,celColNo)
                                ws.cell(celRowNo+1,celColNo).value=cIp.get_ip4_dhcp_end(ipv4Address,ipV4Cidr)   #in networkParameterValue give start address from ip library
                                Ipv6_writtenCell=ws.cell(celRowNo+1,celColNo)
                                ws.cell(celRowNo+2,celColNo).value=cIp.calculate_ipv6_address_End(ipv6Address,ipV6Cidr)   #in networkParameterValue give start address from ip library
                         elif str(cell.value).strip()=="gateway-address":  
                                celRowNo=cell.row
                                celColNo=cell.column
                                print(ws.cell(celRowNo,celColNo+1).value)
                                Ipv4_writtenCell=ws.cell(celRowNo+1,celColNo)
                                ws.cell(celRowNo+1,celColNo).value=cIp.get_ip4_gateway_address(ipv4Address,ipV4Cidr)   #in networkParameterValue give start address from ip library
                                Ipv6_writtenCell=ws.cell(celRowNo+1,celColNo)
                                ws.cell(celRowNo+2,celColNo).value=cIp.Calculate_IpV6_Gateway(ipv6Address,ipV6Cidr)  #in networkParameterValue give start address from ip library       
                                
                           # writtenCell="sudhansu"
                     print(str(cell.value).strip())
                     print("value found")
       
             wb.save(excelPathFull)   
            
             print(maxRow)  
    def Update_HostRoutesSheet(self,filePath,networkParameterValue,labname):
                 excelPathFull=filePath                                     #os.path.join(excelPath,fileName,searchValue)
                 Sheetname='Host-Routes'
                 wb=op.load_workbook(excelPathFull,read_only=False,keep_vba='KEEP_VBA')
                 ws=wb[Sheetname]
                 maxRow=ws.max_row
                 for row in ws.iter_rows():
                     for cell in row:
                         if str(cell.value).strip()=="Host-Routes":
                            celRowNo=cell.row
                            celColNo=cell.column
                            #print(ws.cell(celRowNo,celColNo+1).value)
                            writtenCell=ws.cell(celRowNo,celColNo+1)
                            ws.cell(celRowNo,celColNo+1).value="SubnetNameValue"
                            writtenCell=ws.cell(celRowNo,celColNo+1)
                            ws.cell(celRowNo,celColNo+2).value="SubnetNameValueNextHop"
                           # writtenCell="sudhansu"
                            print(str(cell.value).strip())
                            print("value found")      
                 wb.save(excelPathFull)   
                 print(maxRow)            
                                         
    def Update_HostRoutesSheet_VF(filePath,SubnetNameValue,host_routes_route_prefix,host_routes_route_next_hop,row_num):
                 excelPathFull=filePath                                     #os.path.join(excelPath,fileName,searchValue)
                 Sheetname='Host-Routes'
                 wb=op.load_workbook(excelPathFull,read_only=False,keep_vba='KEEP_VBA')
                 ws=wb[Sheetname]
                 maxRow=ws.max_row
                 for row in ws.iter_rows():
                     for cell in row:
                         if str(cell.value).strip()=="Subnet-Name":
                            celRowNo=cell.row
                            celColNo=cell.column
                            
                            ws.cell(celRowNo+row_num,celColNo).value=SubnetNameValue
                            writtenCell=ws.cell(celRowNo,celColNo+row_num)
                            ws.cell(celRowNo+row_num,celColNo+1).value=host_routes_route_prefix
                            ws.cell(celRowNo+row_num,celColNo+2).value=host_routes_route_next_hop
                            print(str(cell.value).strip())
                            print("value found")
                 wb.save(excelPathFull)
    def create_Network_Preload_files(self,heatfile,rolefile,vnfName,path,labname):
        #Network-type=collections.namedtuple(values[network-name,network-type,start-address,dhcp-start,dhcp-end,gateway-address,Subnet-Name,Subnet-NameNext)
        excelSheets=['General','Subnets','Host-Routes']                                           
        changingParameters=['network-name','network-type','start-address','dhcp-start','dhcp-end','gateway-address','Subnet-Name','Subnet-NameNext'] 
        networkRoles=self.loop_over_heatFiles(heatfile,rolefile)
        print(networkRoles)
        #if not os.path.exists(join(path,vnfName+"//NetworkPreloads")) : 
        preloadPath=join(path,vnfName+"//NetworkPreloads")  
        os.makedirs(preloadPath,mode=777, exist_ok="True",)
        for role in networkRoles:
            print(role)
            copiedFile=preloadPath+"//"+role+"_net.xlsm"
            shutil.copyfile(join(rolefile,"NetworkPreloadTemplate.xlsm"), copiedFile)       
            self.Update_GeneralSheet(copiedFile,role,labname)    
 
 ####### Used For VF:-                          
    def Update_ColumnVal_VF(self,filePath,sheetName,columnName,row_num,rowValue):
                excelPathFull=filePath
                wb=op.load_workbook(filePath,read_only=False,keep_vba='KEEP_VBA')
                ws=wb[sheetName]
                for row in ws.iter_rows():
                    for cell in row:
                        if str(cell.value).strip()==columnName:
                           celRowNo=cell.row
                           celColNo=cell.column
                           ws.cell(row=celRowNo+row_num,column=celColNo).value=rowValue
                           print("value found")
                wb.save(excelPathFull)

    def copyPreloadExcel(self,heatfile,rolefile,vnfName,path,labname):
        preloadPath=join(path,vnfName+"//NetworkPreloads")  
        os.makedirs(preloadPath,mode=777, exist_ok="True",)
        copiedFile=preloadPath+"//"+role+"_net.xlsm"
        shutil.copyfile(join(rolefile,"NetworkPreloadTemplate.xlsm"), copiedFile)       
    
    def Update_RowVal_VF(self,filePath,sheetName,rowName,columnNum,rowValue):
                excelPathFull=filePath
                wb=op.load_workbook(filePath,read_only=False,keep_vba='KEEP_VBA')
                ws=wb[sheetName]
                for row in ws.iter_rows():
                    for cell in row:
                        if str(cell.value).strip()==rowName:
                           celRowNo=cell.row
                           celColNo=cell.column
                           ws.cell(row=celRowNo,column=celColNo+columnNum).value=rowValue
                           print("value found")
                wb.save(excelPathFull)